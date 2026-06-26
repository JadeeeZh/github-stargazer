"""Segment + score enriched stargazers into a CRM (CSV always, XLSX if openpyxl).

Scoring/segmentation are pure functions (no openpyxl) so other steps — e.g. the
deepline LinkedIn resolver — can reuse the priority logic.
"""
import csv, json, os, re
from collections import Counter
from . import config, seed
# NOTE: `icp` is imported lazily inside the GTM functions to avoid a circular import
# (icp imports crm for its pure scoring helpers).

VC = ["ventures", "venture", "capital", " vc", " fund", "partners", "a16z", "sequoia", "accel"]
BIGTECH = ["google", "youtube", "meta", "amazon", "aws", "microsoft", "apple", "nvidia", "paypal",
           "cisco", "tiktok", "bytedance", "accenture", "zscaler", "ibm", "adobe", "salesforce",
           "uber", "oracle", "intel", "netflix", "stripe", "databricks", "snowflake", "intuit",
           "vmware", "samsung", "tencent", "alibaba", "huawei", "baidu", "openai", "anthropic"]
UNI = ["universit", "college", "institute of technology", "school", "academia", "phd",
       "research", "学院", "大学", "iit", "mit", "stanford", "berkeley", "tsinghua", "sjtu",
       "zhejiang", "carnegie", "cmu", "polytechnic", "univ"]


def disp_company(r):
    return (r.get("linkedin_company") or r.get("company") or "").strip()


def _blob(r):
    return " ".join([r.get("github_name", ""), r.get("company", ""), r.get("bio", ""),
                     r.get("linkedin_headline", ""), r.get("linkedin_title", ""),
                     r.get("linkedin_company", ""),
                     " ".join(r.get("top_languages", []) or []),
                     " ".join(r.get("notable_repos", []) or [])]).lower()


def segment(r):
    t = _blob(r)
    co = (r.get("company", "") + " " + r.get("bio", "")).lower()
    if re.search(r"investor|venture|capital|\bvc\b|partner\b|principal|angel", t) and any(k in co for k in VC):
        return "Investor / VC"
    if re.search(r"founder|co-?found|\bceo\b|\bcto\b|\bcoo\b|chief|president|managing director", t):
        return "Founder / Exec"
    if re.search(r"product manager|product lead|head of product|\bpm\b|product owner", t):
        return "Product"
    if re.search(r"\bai\b|\bml\b|machine learning|deep learning|\bllm\b|agent|rag\b|reinforcement|"
                 r"data scien|research scientist|applied scien|nlp|neural", t):
        return "AI / ML Engineer & Researcher"
    if re.search(r"engineer|developer|swe|backend|frontend|full.?stack|software|programmer|devops|"
                 r"sde|coding|hacker", t):
        return "Software Engineer"
    if re.search(r"design|ux|\bui\b|artist|creative", t):
        return "Design / Creative"
    if re.search(r"student|intern|undergrad|graduate|\bms\b|\bphd\b|university|master", t):
        return "Student / Researcher"
    return "Other / Unclear"


def seniority(r):
    t = _blob(r)
    if re.search(r"founder|co-?found|\bceo\b|\bcto\b|\bcoo\b|\bvp\b|vice president|director|"
                 r"head of|principal|distinguished|partner\b|lead\b", t):
        return "Leadership"
    if re.search(r"senior|staff|\bsr\.?\b|architect", t):
        return "Senior"
    if re.search(r"student|intern|undergrad|new grad|junior", t):
        return "Junior / Student"
    return "Mid / Unknown"


def company_type(r):
    c = disp_company(r).lower()
    if not c.strip():
        return ""
    if any(k in c for k in BIGTECH):
        return "Big Tech / Enterprise"
    if any(k in c for k in VC):
        return "VC / Investor firm"
    if any(k in c for k in UNI):
        return "University / Research"
    return "Startup / SMB"


def focus(r):
    t = _blob(r)
    if re.search(r"agent|\bllm\b|genai|gen ai|\brag\b|prompt|memory|copilot|chatbot|mcp", t):
        return "AI agents / LLM"
    if re.search(r"machine learning|\bml\b|deep learning|reinforcement|data scien|vision|nlp|model", t):
        return "ML / Data science"
    if re.search(r"web3|crypto|blockchain|solana|defi|token|onchain|ethereum", t):
        return "Web3 / Crypto"
    if re.search(r"backend|infra|devops|platform|distributed|cloud|kubernetes|security|database", t):
        return "Infra / Backend"
    if re.search(r"frontend|full.?stack|\bweb\b|mobile|ios|android|react", t):
        return "App / Frontend"
    return ""


def lead_type(seg):
    if seg == "Investor / VC":
        return "Investor"
    if seg in ("Founder / Exec", "AI / ML Engineer & Researcher", "Software Engineer", "Product"):
        return "Customer prospect"
    if seg == "Student / Researcher":
        return "Community / Talent"
    return "Other"


def priority(r, seg, lt, foc):
    identifiable = bool(r.get("linkedin_url") or r.get("website") or disp_company(r))
    ai_fit = foc in ("AI agents / LLM", "ML / Data science")
    recent = r.get("rank", 9999) <= 100
    if lt == "Customer prospect" and r.get("linkedin_confidence") == "high" and ai_fit:
        return "A"
    if lt == "Investor" and identifiable:
        return "A"
    if lt == "Customer prospect" and seg in ("Founder / Exec", "AI / ML Engineer & Researcher") \
            and ai_fit and identifiable:
        return "A"
    if lt == "Customer prospect" and identifiable and (ai_fit or recent):
        return "B"
    if identifiable:
        return "B"
    return "C"


def _listjoin(v):
    return "; ".join(str(x).strip() for x in v) if isinstance(v, list) else (v or "")


def build_leads(records):
    """Return (leads, sparse) as lists of dicts. Pure — no file IO."""
    leads, sparse = [], []
    for r in records:
        name = (r.get("github_name") or r.get("login") or "").strip() or r.get("login", "")
        if r.get("profile_status") != "ok" and not (r.get("company") or r.get("bio") or r.get("linkedin_url")):
            sparse.append({"rank": r.get("rank", ""), "login": r.get("login", ""), "name": name,
                           "github": r.get("github_url", ""), "location": r.get("location", ""),
                           "languages": _listjoin(r.get("top_languages")),
                           "profile_status": r.get("profile_status", "")})
            continue
        seg = segment(r); lt = lead_type(seg); foc = focus(r)
        leads.append({
            "rank": r.get("rank", ""), "name": (r.get("linkedin_name") or name),
            "login": r.get("login", ""), "lead_type": lt, "segment": seg,
            "priority": priority(r, seg, lt, foc), "title": r.get("linkedin_title", ""),
            "company": disp_company(r), "seniority": seniority(r),
            "company_type": company_type(r), "focus": foc, "location": r.get("location", ""),
            "linkedin_status": ("verified" if r.get("linkedin_confidence") == "high"
                                else "found" if r.get("linkedin_url") else "not listed on GitHub"),
            "linkedin": r.get("linkedin_url", ""), "website": r.get("website", ""),
            "twitter": r.get("twitter", ""), "email": r.get("email", ""),
            "github": r.get("github_url", ""), "followers": r.get("followers", ""),
            "bio": r.get("linkedin_headline") or r.get("bio", ""),
            "top_languages": _listjoin(r.get("top_languages")),
            "notable_repos": _listjoin(r.get("notable_repos")),
        })
    PRIO = {"A": 0, "B": 1, "C": 2}
    LT = {"Customer prospect": 0, "Investor": 1, "Community / Talent": 2, "Other": 3}
    leads.sort(key=lambda x: (PRIO[x["priority"]], LT[x["lead_type"]],
                              x["rank"] if isinstance(x["rank"], int) else 1e9))
    return leads, sparse


LEAD_COLS = ["rank", "name", "login", "lead_type", "segment", "priority", "title", "company",
             "seniority", "company_type", "focus", "location", "linkedin_status",
             "linkedin", "website", "twitter", "email", "github", "followers",
             "bio", "top_languages", "notable_repos"]


# `build_leads` (above) + the pure scoring fns are reused by build_gtm and deepline_lookup.
# The legacy star-only CRM writer was removed when the GTM workflow superseded it.

# ======================================================================================
# GTM workflow (Step 3): ICP-bucketed CRM. `gtm_grade` is a bucket-relative grade,
# distinct from the absolute `priority` column.
# ======================================================================================
GTM_COLS = LEAD_COLS + ["engaged_at", "bucket", "source", "fork_rank", "star_rank", "icp_fit", "gtm_grade"]
GRADE_ORD = {"A": 0, "B": 1, "C": 2}

# Research-verdict columns, joined onto B-end rows from work/icp_verdicts.json (if present).
VERDICT_COLS = ["icp_verdict", "memory_value", "why_fit", "lead_with",
                "verdict_confidence", "contact", "next_action"]
BEND_COLS = GTM_COLS + VERDICT_COLS
# rank verdicts so researched targets rise to the top of the B-end sheet
VERDICT_ORD = {"Tier A": 0, "Yellow flag": 1, "Channel": 2, "Supplier": 3, "Scout": 4,
               "Unknown": 5, "": 6, "Disqualified": 7, "Not a company": 8}
_NEXT_ACTION = {
    "Tier A": "Outreach now + tailored one-pager",
    "Yellow flag": "Qualify ICP fit on first call",
    "Channel": "Partnership / referral intro",
    "Supplier": "Vendor — co-marketing only",
    "Scout": "Competitor intel — do not sell",
    "Disqualified": "Skip",
    "Not a company": "Skip",
    "Unknown": "Revisit if a real company materializes",
}


def _vnorm(c):
    return re.sub(r"[^\w ]", "", (c or "").lower()).strip()


def load_verdicts(icp_id):
    """{_vnorm(company): verdict} for one ICP, from .cache/verdicts/<icp_id>.json. {} if absent."""
    try:
        return json.load(open(config.verdicts_file(icp_id), encoding="utf-8"))
    except Exception:
        return {}


def _attach_verdict(row, verdicts):
    """Add the 7 VERDICT_COLS to a B-end row from a matching researched verdict (or blanks)."""
    v = verdicts.get(_vnorm(row.get("company", "")))
    if not v:
        for k in VERDICT_COLS:
            row[k] = ""
        return row
    vd = v.get("icp_verdict", "")
    contact = " · ".join(x for x in (v.get("rep_name", ""), v.get("rep_github", ""),
                                     v.get("linkedin", "")) if x)
    row.update({
        "icp_verdict": vd, "memory_value": v.get("memory_value", ""),
        "why_fit": v.get("why_fit", ""), "lead_with": v.get("lead_with", ""),
        "verdict_confidence": v.get("verdict_confidence", ""), "contact": contact,
        "next_action": _NEXT_ACTION.get(vd, ""),
    })
    return row


def build_gtm(records, icp, slug, verdicts=None):
    """Return (b_leads, c_leads, sparse) for one repo. Pure apart from reading this repo's seeds.
    match/bucket/score are computed exactly once per record; verdicts (a {_vnorm:rec} dict) are
    joined onto B-end rows."""
    from . import icp as icpmod
    if icp is None:
        icp = icpmod.ICP()  # neutral matcher
    verdicts = verdicts or {}
    seeds = seed.load_seeds(slug)
    b, c, sparse = [], [], []
    for r in records:
        login = r.get("login", "")
        name = (r.get("github_name") or login or "").strip() or login
        m = icpmod.match(r, icp)
        bkt = icpmod.bucket(r, m)
        src = seed.source_of(login, seeds)
        is_sparse = (r.get("profile_status") != "ok"
                     and not (icpmod.has_real_company(r) or r.get("bio") or r.get("linkedin_url")))
        if bkt == "neither" or is_sparse:
            sparse.append({"rank": src["best_rank"], "login": login, "name": name,
                           "github": r.get("github_url", ""), "location": r.get("location", ""),
                           "languages": _listjoin(r.get("top_languages")),
                           "profile_status": r.get("profile_status", "")})
            continue
        seg = segment(r); lt = lead_type(seg); foc = focus(r)
        sc = icpmod.score(r, icp, src, bkt, m)
        row = {
            "rank": src["best_rank"], "engaged_at": src.get("engaged_at", ""),
            "name": (r.get("linkedin_name") or name),
            "login": login, "lead_type": lt, "segment": seg,
            "priority": priority(r, seg, lt, foc),  # legacy priority preserved as-is
            "title": r.get("linkedin_title", ""), "company": disp_company(r),
            "seniority": seniority(r), "company_type": company_type(r), "focus": foc,
            "location": r.get("location", ""),
            "linkedin_status": ("verified" if r.get("linkedin_confidence") == "high"
                                else "found" if r.get("linkedin_url") else "not listed on GitHub"),
            "linkedin": r.get("linkedin_url", ""), "website": r.get("website", ""),
            "twitter": r.get("twitter", ""), "email": r.get("email", ""),
            "github": r.get("github_url", ""), "followers": r.get("followers", ""),
            "bio": r.get("linkedin_headline") or r.get("bio", ""),
            "top_languages": _listjoin(r.get("top_languages")),
            "notable_repos": _listjoin(r.get("notable_repos")),
            "bucket": bkt, "source": src["source"],
            "fork_rank": src["fork_rank"] if src["fork_rank"] is not None else "",
            "star_rank": src["star_rank"] if src["star_rank"] is not None else "",
            "icp_fit": (m["b_hits"] if bkt == "B" else m["c_hits"]),
            "gtm_grade": icpmod.grade(sc, m["excluded"]),
            "_score": sc, "best_rank": src["best_rank"], "_excluded": m["excluded"],
        }
        (b if bkt == "B" else c).append(row)
    # join research verdicts onto B-end rows (blank columns if none researched yet)
    for row in b:
        _attach_verdict(row, verdicts)
    # B-end: researched verdict tier first (Tier A -> ... -> DQ), then grade/score/source/recency
    b.sort(key=lambda x: (VERDICT_ORD.get(x.get("icp_verdict", ""), 6), GRADE_ORD[x["gtm_grade"]],
                          -x["_score"], seed.SOURCE_ORDER[x["source"]],
                          x["best_rank"] if isinstance(x["best_rank"], int) else 1e9, x["login"]))
    c.sort(key=lambda x: (GRADE_ORD[x["gtm_grade"]], -x["_score"], seed.SOURCE_ORDER[x["source"]],
                          x["best_rank"] if isinstance(x["best_rank"], int) else 1e9, x["login"]))
    return b, c, sparse


# Legend for the evaluative/derived (non-factual) columns — shipped as an Explanation
# sheet in the workbook + explanation.csv. ("H", title) = section header; ("F", field, meaning, values).
LEGEND = [
    ("H", "READ FIRST — two grades, don't confuse them"),
    ("F", "gtm_grade", "AUTO keyword score (GitHub activity + AI/ICP terms). NOT whether it's a buyable fit; NOT a gate.", "A (score>=9) / B (>=5) / C"),
    ("F", "icp_verdict", "The RESEARCHED judgment (web research per your ICP). Trust this over gtm_grade.", "see 'Researched verdict' below; blank = not researched"),
    ("F", "engaged_at", "The REAL date they forked/starred (fork date preferred). Use this, not best_rank. Blank for star-only fetched without a token (HTML has no star dates).", "YYYY-MM-DD"),
    ("F", "rank / best_rank", "Internal recency RANK (1 = newest) used only for sorting — read engaged_at for the actual date.", "1..N"),
    ("H", "Engagement & recency"),
    ("F", "source", "How they engaged. fork = took the code (stronger intent); both = strongest.", "star / fork / both"),
    ("F", "fork_rank / star_rank", "Position in each list (blank if not in that list).", "1..N or blank"),
    ("H", "Bucket & auto tags (derived from the GitHub profile)"),
    ("F", "bucket", "B = has a real company; C = individual builder (no real company).", "B / C"),
    ("F", "lead_type / segment", "Role classification.", "Founder/Exec, AI/ML, Software Eng, Product, ..."),
    ("F", "seniority / company_type / focus", "Derived tags from profile + bio (+ LinkedIn if scraped).", ""),
    ("F", "linkedin_status", "verified = logged-in scrape; found = a URL we have; else not listed.", "verified / found / not listed on GitHub"),
    ("F", "priority", "Legacy A/B/C (identifiability + recency). Secondary to icp_verdict.", "A / B / C"),
    ("F", "icp_fit", "Number of ICP keywords matched.", "0..N"),
    ("H", "B-end company gate (free pre-filter — 'prove it's a real company')"),
    ("F", "category", "Company TYPE triage on the name.", "candidate / academic / bigtech / junk"),
    ("F", "company_status", "How much free evidence it's a real company.", "proven (>=2) / to-verify (1) / unproven (0)"),
    ("F", "evidence", "Count of proofs: real domain / /company/ LinkedIn / non-free email / >=2 leads / domain-like name.", "0..5"),
    ("F", "research_status", "What to do next. Filter 'research' for the next batch.", "research / done / auto-dq / skip"),
    ("H", "Researched verdict (web research per ICP — the real B-end judgment)"),
    ("F", "icp_verdict", "Direct-sales verdict.", "Tier A / Yellow flag / Channel / Supplier / Scout / Disqualified / Not a company / Unknown"),
    ("F", "memory_value", "Which memory type the fit hangs on (User = remembers the person; Agent = gets better with use; Shared = team knowledge).", "User / Agent / Shared / User+Agent / none"),
    ("F", "what_they_do / why_fit / lead_with", "One-line description / why it fits the ICP / the pain to open outreach with.", ""),
    ("F", "next_action / verdict_confidence / contact", "Suggested move / research confidence / best contact.", ""),
    ("H", "C-end shortlist extras (cend_shortlist.csv)"),
    ("F", "tried", "Forked = actually ran it -> best user-interview subject.", "yes / blank"),
    ("F", "reachable", "Has a non-GitHub contact (LinkedIn/X/email/site).", "yes / github-only"),
]


def _write_explanation_sheet(wb):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("Explanation")
    hfill = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
    ws.append(["Field", "What it means", "Values"])
    for i in range(1, 4):
        ws.cell(1, i).fill = hfill; ws.cell(1, i).font = hfont
    for row in LEGEND:
        if row[0] == "H":
            ws.append([row[1], "", ""])
            c = ws.cell(ws.max_row, 1); c.font = Font(bold=True, color="1F4E78")
        else:
            ws.append([row[1], row[2], row[3]])
    for col, w in (("A", 34), ("B", 82), ("C", 46)):
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


_GTM_W = {"engaged_at": 12, "bucket": 8, "source": 10, "fork_rank": 9, "star_rank": 9,
          "icp_fit": 8, "gtm_grade": 9}


def _write_gtm_xlsx(b, c, sparse, icp, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (openpyxl not installed -> skipping .xlsx; CSVs still written.)")
        return False
    W = {"rank": 6, "name": 22, "login": 18, "lead_type": 17, "segment": 28, "priority": 8,
         "title": 28, "company": 26, "seniority": 15, "company_type": 20, "focus": 20,
         "location": 20, "linkedin_status": 18, "linkedin": 40, "website": 30, "twitter": 26,
         "email": 26, "github": 32, "followers": 10, "bio": 50, "top_languages": 26,
         "notable_repos": 40}
    W.update(_GTM_W)
    hfill = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
    grade_fill = {"A": PatternFill("solid", fgColor="C6EFCE"),
                  "B": PatternFill("solid", fgColor="FFEB9C"),
                  "C": PatternFill("solid", fgColor="F2F2F2")}
    W.update({"icp_verdict": 14, "memory_value": 12, "why_fit": 50, "lead_with": 50,
              "verdict_confidence": 11, "contact": 46, "next_action": 28})
    wb = Workbook()

    def lead_sheet(title, rows, cols):
        ws = wb.create_sheet(title)
        gidx = cols.index("gtm_grade") + 1
        ws.append([col.replace("_", " ").title() for col in cols])
        for i in range(1, len(cols) + 1):
            ws.cell(1, i).fill = hfill; ws.cell(1, i).font = hfont
            ws.cell(1, i).alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(col, "") for col in cols])
            ws.cell(ws.max_row, gidx).fill = grade_fill.get(r.get("gtm_grade", "C"),
                                                            grade_fill["C"])
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = W.get(col, 16)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    wb.remove(wb.active)
    lead_sheet("B-end target companies", b, BEND_COLS)
    _write_explanation_sheet(wb)  # 2nd tab: legend for the evaluative/derived fields
    lead_sheet("C-end high-potential", c, GTM_COLS)

    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 32; ss.column_dimensions["B"].width = 10
    ss.column_dimensions["D"].width = 32; ss.column_dimensions["E"].width = 10

    def block(col_label, col_val, title, counter, start):
        ss.cell(start, col_label, title).font = Font(bold=True)
        row = start + 1
        for k, v in counter.most_common():
            ss.cell(row, col_label, k or "(blank)"); ss.cell(row, col_val, v); row += 1
        return row + 1

    rB = block(1, 2, "B-end · count", Counter({"leads": len(b)}), 1)
    rB = block(1, 2, "B-end · by grade", Counter(x["gtm_grade"] for x in b), rB)
    rB = block(1, 2, "B-end · by company type", Counter(x["company_type"] for x in b if x["company_type"]), rB)
    rB = block(1, 2, "B-end · by seniority", Counter(x["seniority"] for x in b), rB)
    rB = block(1, 2, "B-end · by source", Counter(x["source"] for x in b), rB)
    ss.cell(rB, 1, "B-end · verified LinkedIn").font = Font(bold=True)
    ss.cell(rB, 2, sum(1 for x in b if x["linkedin_status"] == "verified"))

    rC = block(4, 5, "C-end · count", Counter({"leads": len(c)}), 1)
    rC = block(4, 5, "C-end · by grade", Counter(x["gtm_grade"] for x in c), rC)
    rC = block(4, 5, "C-end · by focus", Counter(x["focus"] for x in c if x["focus"]), rC)
    rC = block(4, 5, "C-end · by source", Counter(x["source"] for x in c), rC)
    ss.cell(rC, 4, "C-end · identifiable (LI/site)").font = Font(bold=True)
    ss.cell(rC, 5, sum(1 for x in c if x["linkedin"] or x["website"]))

    base = max(rB, rC) + 1
    ss.cell(base, 1, "Cross-cut").font = Font(bold=True)
    allrows = b + c
    ss.cell(base + 1, 1, "total leads (B+C)"); ss.cell(base + 1, 2, len(allrows))
    ss.cell(base + 2, 1, "fork-only"); ss.cell(base + 2, 2, sum(1 for x in allrows if x["source"] == "fork"))
    ss.cell(base + 3, 1, "star-only"); ss.cell(base + 3, 2, sum(1 for x in allrows if x["source"] == "star"))
    ss.cell(base + 4, 1, "both (forked+starred)"); ss.cell(base + 4, 2, sum(1 for x in allrows if x["source"] == "both"))
    ss.cell(base + 5, 1, "ICP loaded"); ss.cell(base + 5, 2, "yes" if icp.loaded else "no (neutral)")
    ss.cell(base + 6, 1, "grade method"); ss.cell(base + 6, 2, f"absolute (A>={icpmod_floor_a()}, B>={icpmod_floor_b()})")

    sp = wb.create_sheet("Sparse profiles")
    SP = ["rank", "login", "name", "github", "location", "languages", "profile_status"]
    sp.append([col.replace("_", " ").title() for col in SP])
    for i in range(1, len(SP) + 1):
        sp.cell(1, i).fill = hfill; sp.cell(1, i).font = hfont
    for r in sorted(sparse, key=lambda x: x["rank"] if isinstance(x["rank"], int) else 1e9):
        sp.append([r[col] for col in SP])
    for i, col in enumerate(SP, 1):
        sp.column_dimensions[get_column_letter(i)].width = {"rank": 6, "login": 20, "name": 22,
            "github": 34, "location": 20, "languages": 28, "profile_status": 14}[col]
    sp.freeze_panes = "A2"
    wb.save(path)
    return True


def icpmod_floor_a():
    from . import icp as icpmod
    return icpmod.GRADE_A_FLOOR


def icpmod_floor_b():
    from . import icp as icpmod
    return icpmod.GRADE_B_FLOOR


COMPANY_FEED_COLS = ["company", "research_status", "icp_verdict", "category", "company_status",
                     "evidence", "memory_value", "best_grade", "lead_count", "engaged_at",
                     "best_rank", "sources", "what_they_do", "lead_with", "next_action"]
_STATUS_ORD = {"proven": 0, "to-verify": 1, "unproven": 2}
# research_status = the single "what to do next" gate (kills manual shortlist rebuilding):
#   done     -> already has an icp_verdict
#   auto-dq  -> academic/bigtech/junk type => non-customer per ICP, don't spend research
#   research -> candidate company worth web-research (has evidence OR is AI-focused)
#   skip     -> candidate but unproven AND not AI-focused => the junk long tail, skip
_RESEARCH_ORD = {"research": 0, "done": 1, "auto-dq": 2, "skip": 3}


def _ai_focused(leads):
    return any((x.get("focus") or "").startswith(("AI", "ML")) for x in leads)


def _research_status(icp_verdict, category, company_status, ai_focused):
    if icp_verdict:
        return "done"
    if category in ("academic", "bigtech", "junk"):
        return "auto-dq"
    if company_status in ("proven", "to-verify") or ai_focused:
        return "research"
    return "skip"


def _company_feed(b, verdicts):
    """Dedup B rows by _vnorm(company) -> one row per company (THE B-end deliverable).

    Survivor = the lead with the lowest (best/newest) best_rank; grade/sources aggregate the
    company's leads; verdict fields come from the join (what_they_do from the verdict dict).
    `company_status` (proven/to-verify/unproven) records the free 'is it a real company?'
    evidence — the cheap gate so expensive research can skip unproven names.
    """
    from . import icp as icpmod
    groups = {}
    for r in b:
        k = _vnorm(r.get("company", ""))
        if k:
            groups.setdefault(k, []).append(r)
    rows = []
    for k, leads in groups.items():
        survivor = min(leads, key=lambda r: r["best_rank"] if isinstance(r.get("best_rank"), int) else 10 ** 9)
        best_grade = min((x["gtm_grade"] for x in leads), key=lambda g: GRADE_ORD[g])
        srcs = "+".join(sorted({x["source"] for x in leads}, key=lambda s: seed.SOURCE_ORDER.get(s, 9)))
        ev = icpmod.company_evidence(leads)
        cstatus = icpmod.evidence_status(ev)
        cat = icpmod.company_category(survivor.get("company", ""))
        verdict = survivor.get("icp_verdict", "")
        rstatus = _research_status(verdict, cat, cstatus, _ai_focused(leads))
        v = verdicts.get(k, {})
        rows.append({
            "company": survivor.get("company", ""),
            "research_status": rstatus,
            "icp_verdict": verdict,
            "category": cat,
            "company_status": cstatus,
            "evidence": ev,
            "memory_value": survivor.get("memory_value", ""),
            "best_grade": best_grade,
            "lead_count": len(leads),
            "engaged_at": survivor.get("engaged_at", ""),
            "best_rank": survivor["best_rank"] if isinstance(survivor.get("best_rank"), int) else "",
            "sources": srcs,
            "what_they_do": v.get("what_they_do", ""),
            "lead_with": survivor.get("lead_with", ""),
            "next_action": survivor.get("next_action", ""),
        })
    # done/research first (verdict tier, then evidence), then auto-dq/skip last; best_grade is
    # NOT a gate or sort key (it tracks GitHub activity, not ICP fit).
    rows.sort(key=lambda x: (_RESEARCH_ORD[x["research_status"]], VERDICT_ORD.get(x["icp_verdict"], 6),
                             _STATUS_ORD[x["company_status"]],
                             x["best_rank"] if isinstance(x["best_rank"], int) else 1e9))
    return rows


CEND_SHORTLIST_COLS = ["rank", "engaged_at", "name", "login", "source", "gtm_grade", "focus",
                       "followers", "tried", "reachable", "linkedin", "twitter", "email", "website",
                       "github", "bio", "notable_repos"]


def _cend_shortlist(c):
    """C-end builders ranked for user interviews / developer upsell: who actually tried the repo
    (forked) + is an AI builder + is reachable. Forkers rank first (their feedback is gold)."""
    def reach(r):
        return bool(r.get("linkedin") or r.get("twitter") or r.get("email") or r.get("website"))

    def tried(r):
        return r.get("source") in ("fork", "both")

    def builder(r):
        return (r.get("focus") or "").startswith(("AI", "ML")) or r.get("gtm_grade") == "A"

    picked = [r for r in c if builder(r)]
    picked.sort(key=lambda r: (0 if tried(r) else 1, GRADE_ORD[r["gtm_grade"]],
                               0 if reach(r) else 1,
                               r.get("best_rank") if isinstance(r.get("best_rank"), int) else 1e9))
    out = []
    for r in picked:
        d = {k: r.get(k, "") for k in CEND_SHORTLIST_COLS}
        d["tried"] = "yes" if tried(r) else ""
        d["reachable"] = "yes" if reach(r) else "github-only"
        out.append(d)
    return out


def _write_manifest(slug, icp_id, icp, b, c, sparse, feed, scraped_this_run, path, cend_count=0):
    import datetime
    import hashlib
    rp = config.repo_paths(slug)
    repo_meta = {}
    if os.path.exists(rp["repo_json"]):
        try:
            repo_meta = json.load(open(rp["repo_json"], encoding="utf-8"))
        except Exception:
            repo_meta = {}
    sha = ""
    try:
        sha = hashlib.sha256(open(icp.path, "rb").read()).hexdigest() if icp.path else ""
    except Exception:
        sha = ""
    joined = [x for x in b if x.get("icp_verdict")]
    manifest = {
        "schema_version": 1,
        "run_id": f"{config.repo_key(slug)}@{icp_id}",
        "repo": slug, "repo_key": config.repo_key(slug),
        "icp": {"id": icp_id, "path": icp.path, "sha256": sha, "loaded": icp.loaded,
                "warnings": icp.warnings},
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "tool_version": config.tool_version(),
        "seeds": {"last_enriched_at": repo_meta.get("last_enriched_at", ""),
                  "stars": repo_meta.get("stars", ""), "forks": repo_meta.get("forks", "")},
        "cache": {"profiles_scraped_this_run": scraped_this_run, "cache_dir": ".cache/profiles"},
        "counts": {"b_leads": len(b), "b_companies_distinct": len(feed), "c_leads": len(c),
                   "sparse_dropped": len(sparse),
                   "b_grade": dict(Counter(x["gtm_grade"] for x in b)),
                   "c_grade": dict(Counter(x["gtm_grade"] for x in c)),
                   "source_mix": dict(Counter(x["source"] for x in (b + c)))},
        "research": {"verdicts_joined": len(joined),
                     "verdict_mix": dict(Counter(x["icp_verdict"] for x in joined)),
                     "verdicts_file": os.path.relpath(config.verdicts_file(icp_id), config.ROOT)},
        "files": [
            {"name": "bend_company_feed.csv", "role": "SEND to your company-enrichment / ICP research pass", "rows": len(feed)},
            {"name": "bend_companies.csv", "role": "B-end leads (per person) + verdicts", "rows": len(b)},
            {"name": "cend_individuals.csv", "role": "C-end AI-builder individuals + contacts", "rows": len(c)},
            {"name": "cend_shortlist.csv", "role": "C-end builders ranked for user interviews / upsell (forkers first)", "rows": cend_count},
            {"name": "linkedin_to_scrape.csv", "role": "SEND to LinkedIn_Profile_Scraper", "rows": sum(1 for r in (b + c) if r["linkedin"])},
            {"name": "gtm_crm.xlsx", "role": "master B+C workbook", "rows": len(b) + len(c)},
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    return manifest


def _write_readme(m, path):
    L = [f"# {m['repo']} × ICP `{m['icp']['id']}` — GTM run", "",
         f"Generated {m['generated_at']} · tool {m['tool_version']} · run_id `{m['run_id']}`",
         f"ICP loaded: **{m['icp']['loaded']}** ({m['icp']['path']})", ""]
    if m["icp"]["warnings"]:
        L += [f"> ICP warnings: {m['icp']['warnings']}", ""]
    cnt = m["counts"]
    L += ["## Counts",
          f"- **B-end**: {cnt['b_leads']} leads → **{cnt['b_companies_distinct']} distinct companies** · grades {cnt['b_grade']}",
          f"- **C-end**: {cnt['c_leads']} individuals · grades {cnt['c_grade']}",
          f"- sparse/dropped: {cnt['sparse_dropped']} · source mix {cnt['source_mix']}",
          f"- research verdicts joined: {m['research']['verdicts_joined']} {m['research']['verdict_mix']}", "",
          "## Files (what to do with each)"]
    for f in m["files"]:
        L.append(f"- `{f['name']}` ({f['rows']} rows) — {f['role']}")
    L += ["- `research/` — research memo + verdicts (if the ICP qualifier was run)", "",
          "## Two-pass scoring",
          "1. `run` / `score` produces `bend_company_feed.csv` (verdict columns blank on first pass).",
          "2. Research the companies against your ICP → writes `.cache/verdicts/<icp>.json`.",
          f"3. `score {m['repo']} --icp {m['icp']['id']}` re-joins verdicts (zero network).", ""]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")


def run_gtm(slug, icp_id, scraped_this_run=0):
    from . import enrich, icp as icpmod
    config.ensure_run_dirs(slug, icp_id)
    icp_path, icp_id = config.resolve_icp(icp_id)
    # refresh this repo's union snapshot from the shared cache (network-free)
    enrich.rebuild_combined(slug)
    try:
        records = json.load(open(config.repo_paths(slug)["union"], encoding="utf-8"))
    except Exception:
        print(f"no union for {slug}; run `run {slug}` (fetch+enrich) first")
        return [], [], []
    icp = icpmod.load(icp_path)
    verdicts = load_verdicts(icp_id)
    b, c, sparse = build_gtm(records, icp, slug, verdicts)
    feed = _company_feed(b, verdicts)
    paths = config.run_paths(slug, icp_id)

    def emit(rows, cols):
        return [{k: r.get(k, "") for k in cols} for r in rows]

    with open(paths["csv_b"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=BEND_COLS); w.writeheader(); w.writerows(emit(b, BEND_COLS))
    with open(paths["csv_c"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=GTM_COLS); w.writeheader(); w.writerows(emit(c, GTM_COLS))
    cend_short = _cend_shortlist(c)
    with open(paths["csv_cend_shortlist"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CEND_SHORTLIST_COLS); w.writeheader(); w.writerows(cend_short)
    with open(paths["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=BEND_COLS); w.writeheader()
        w.writerows(emit(b, BEND_COLS) + emit(c, BEND_COLS))
    with open(paths["csv_company_feed"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COMPANY_FEED_COLS); w.writeheader(); w.writerows(feed)
    # field legend (also a sheet in the xlsx) for whoever opens the CSVs in Sheets/Excel
    with open(paths["explanation_csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["section", "field", "what_it_means", "values"])
        for row in LEGEND:
            w.writerow([row[1], "", "", ""] if row[0] == "H" else ["", row[1], row[2], row[3]])
    li = [{"profile_url": r["linkedin"], "candidate_id": r["login"], "notes": r["name"]}
          for r in (b + c) if r["linkedin"]]
    with open(paths["linkedin_csv"], "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["profile_url", "candidate_id", "notes"])
        w.writeheader(); w.writerows(li)
    # human view of joined verdicts
    joined = [x for x in b if x.get("icp_verdict")]
    with open(paths["verdicts_csv"], "w", newline="", encoding="utf-8-sig") as f:
        vcols = ["company", "icp_verdict", "memory_value", "verdict_confidence", "why_fit",
                 "lead_with", "contact", "next_action"]
        w = csv.DictWriter(f, fieldnames=vcols, extrasaction="ignore")
        w.writeheader(); w.writerows(joined)
    _write_gtm_xlsx(b, c, sparse, icp, paths["xlsx"])
    manifest = _write_manifest(slug, icp_id, icp, b, c, sparse, feed, scraped_this_run,
                               paths["manifest"], cend_count=len(cend_short))
    _write_readme(manifest, paths["readme"])

    print("=== GTM BUILD COMPLETE ===")
    print(f"B-end: {len(b)} leads / {len(feed)} distinct companies | C-end: {len(c)} | "
          f"sparse: {len(sparse)}")
    print("research_status:", dict(Counter(x["research_status"] for x in feed)),
          f"| C-end interview shortlist: {len(cend_short)}")
    print("B grade:", dict(Counter(x["gtm_grade"] for x in b)),
          "| C grade:", dict(Counter(x["gtm_grade"] for x in c)))
    if joined:
        print(f"verdicts joined: {len(joined)} ->", dict(Counter(x["icp_verdict"] for x in joined)))
    print(f"ICP: {icp_id} (loaded={icp.loaded})" + (f" warnings={icp.warnings}" if icp.warnings else ""))
    print("-> run dir:", paths["dir"])
    return b, c, sparse
